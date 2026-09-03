"""
============================================================================
DOCUMENT CONVERTER ENGINE (converter.py)
Converts DOCX, PDF, XLSX, PPTX, HTML & Text to Clean Structured Markdown
Engines: Microsoft MarkItDown • IBM Docling • Robust Native Fallbacks
============================================================================
"""

import os
import re
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any

logger = logging.getLogger(__name__)

# Try importing Microsoft MarkItDown
try:
    from markitdown import MarkItDown
    MARKITDOWN_AVAILABLE = True
except ImportError:
    MARKITDOWN_AVAILABLE = False
    logger.warning("markitdown is not installed. Will use fallback converters if needed.")

# Try importing IBM Docling
try:
    from docling.document_converter import DocumentConverter
    DOCLING_AVAILABLE = True
except ImportError:
    DOCLING_AVAILABLE = False

class DocumentConverterEngine:
    def __init__(self, prefer_engine: str = "markitdown"):
        self.prefer_engine = prefer_engine
        self.md_client = MarkItDown() if MARKITDOWN_AVAILABLE else None
        self.docling_client = DocumentConverter() if DOCLING_AVAILABLE else None

    def convert_to_markdown(self, local_file_path: str, original_filename: str) -> Dict[str, Any]:
        """
        Main conversion dispatcher.
        Returns a dict with clean_markdown, metadata, word_count, char_count.
        """
        ext = Path(local_file_path).suffix.lower()
        logger.info(f"Parsing document: {original_filename} (Ext: {ext})")

        markdown_body = ""
        engine_used = "fallback"

        # 1. Try Microsoft MarkItDown
        if self.md_client and self.prefer_engine == "markitdown":
            try:
                result = self.md_client.convert(local_file_path)
                markdown_body = result.text_content
                engine_used = "microsoft-markitdown"
                logger.info(f"✓ Converted via MarkItDown: {original_filename}")
            except Exception as e:
                logger.warning(f"MarkItDown failed for {original_filename}: {e}. Trying fallback...")

        # 2. Try IBM Docling if MarkItDown failed or wasn't preferred
        if not markdown_body and self.docling_client:
            try:
                conv_result = self.docling_client.convert(local_file_path)
                markdown_body = conv_result.document.export_to_markdown()
                engine_used = "ibm-docling"
                logger.info(f"✓ Converted via Docling: {original_filename}")
            except Exception as e:
                logger.warning(f"Docling failed for {original_filename}: {e}. Trying fallback...")

        # 3. Fallback Parsers
        if not markdown_body:
            markdown_body = self._fallback_parse(local_file_path, ext)
            engine_used = f"native-fallback-{ext.replace('.', '')}"

        # Clean and normalize Markdown
        clean_md = self._post_process_markdown(markdown_body, original_filename, engine_used)

        word_count = len(clean_md.split())
        char_count = len(clean_md)

        return {
            "markdown": clean_md,
            "engine": engine_used,
            "word_count": word_count,
            "char_count": char_count,
            "original_filename": original_filename
        }

    def _fallback_parse(self, file_path: str, ext: str) -> str:
        """Fallback converter for common file formats."""
        if ext in [".txt", ".md", ".json", ".xml", ".csv"]:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()
            if ext == ".csv":
                return self._csv_to_markdown_table(content)
            elif ext in [".json", ".xml"]:
                return f"```{ext.replace('.', '')}\n{content}\n```"
            return content

        elif ext in [".html", ".htm"]:
            try:
                from bs4 import BeautifulSoup
                with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                    soup = BeautifulSoup(f.read(), "html.parser")
                # Remove scripts and styles
                for s in soup(["script", "style", "nav", "footer"]):
                    s.extract()
                return soup.get_text(separator="\n\n")
            except Exception as e:
                logger.error(f"HTML fallback failed: {e}")

        elif ext == ".pdf":
            try:
                import pypdf
                reader = pypdf.PdfReader(file_path)
                pages_text = []
                for idx, page in enumerate(reader.pages):
                    text = page.extract_text() or ""
                    pages_text.append(f"## Страница {idx + 1}\n\n{text}")
                return "\n\n".join(pages_text)
            except Exception as e:
                logger.error(f"PDF fallback failed: {e}")

        elif ext in [".docx", ".doc"]:
            try:
                import docx
                doc = docx.Document(file_path)
                doc_lines = []
                for p in doc.paragraphs:
                    if p.text.strip():
                        if p.style.name.startswith("Heading 1"):
                            doc_lines.append(f"# {p.text}")
                        elif p.style.name.startswith("Heading 2"):
                            doc_lines.append(f"## {p.text}")
                        elif p.style.name.startswith("Heading 3"):
                            doc_lines.append(f"### {p.text}")
                        else:
                            doc_lines.append(p.text)
                return "\n\n".join(doc_lines)
            except Exception as e:
                logger.error(f"DOCX fallback failed: {e}")

        elif ext in [".xlsx", ".xls"]:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True)
                sheets_md = []
                for sheetname in wb.sheetnames:
                    ws = wb[sheetname]
                    sheets_md.append(f"### Таблица: {sheetname}\n")
                    rows = list(ws.iter_rows(values_only=True))
                    if rows:
                        headers = [str(c or "") for c in rows[0]]
                        sheets_md.append("| " + " | ".join(headers) + " |")
                        sheets_md.append("| " + " | ".join(["---"] * len(headers)) + " |")
                        for r in rows[1:]:
                            if any(r):
                                sheets_md.append("| " + " | ".join([str(c or "") for c in r]) + " |")
                return "\n\n".join(sheets_md)
            except Exception as e:
                logger.error(f"XLSX fallback failed: {e}")

        return f"# Документ: {Path(file_path).name}\n\n*Внимание: Текст не удалось извлечь стандартными парсерами.*"

    def _csv_to_markdown_table(self, csv_content: str) -> str:
        import csv
        import io
        reader = csv.reader(io.StringIO(csv_content))
        lines = list(reader)
        if not lines:
            return ""
        md_lines = []
        headers = [c.strip() for c in lines[0]]
        md_lines.append("| " + " | ".join(headers) + " |")
        md_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
        for r in lines[1:]:
            if any(r):
                md_lines.append("| " + " | ".join([c.strip() for c in r]) + " |")
        return "\n".join(md_lines)

    def _post_process_markdown(self, raw_md: str, source_name: str, engine: str) -> str:
        """Cleans empty lines, normalizes headers and injects metadata block."""
        # Normalize redundant newlines
        cleaned = re.sub(r'\n{3,}', '\n\n', raw_md.strip())
        
        # Remove trailing whitespaces per line
        lines = [l.rstrip() for l in cleaned.splitlines()]
        cleaned = "\n".join(lines)

        # Header metadata block
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        stem_name = Path(source_name).stem

        header = f"""---
source_file: "{source_name}"
converted_at: "{now_str}"
converter_engine: "{engine}"
status: "PROCESSED_CLEAN"
---

# 📄 {stem_name}

"""
        return header + cleaned
